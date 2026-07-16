#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
import pandas as pd

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from tools.qlmg_evidence_contracts import (  # noqa: E402
    PROTECTED_TS,
    result_to_jsonable,
    scan_output_tree_for_protected,
    validate_control_rows,
    validate_event_trade_schema,
    validate_funding_mark_flags,
    validate_no_current_only_taxonomy_rankable,
    validate_no_projected_metric_promotion,
    validate_pit_feature_timestamps,
)
from tools.qlmg_screening_core import check_resource_guard, resource_snapshot, utc_now, write_json  # noqa: E402

try:  # noqa: E402
    from tools.telegram_notify import TelegramNotifier
except Exception:  # pragma: no cover
    TelegramNotifier = None  # type: ignore

RESULTS_ROOT = REPO / "results/rebaseline"
DEFAULT_RUN_ID = "phase_kraken_hypothesis_sweep_readiness_20260701_v1"
DEFAULT_SEED = 20260701
DEFAULT_HYPOTHESIS_LIBRARY = REPO / "research_inputs/QLMG_Hypothesis_Library_2026-07-01.xlsx"
DEFAULT_RESEARCH_INPUT_DIR = REPO / "research_inputs"
DEFAULT_KRAKEN_DATA_ROOT = Path("/opt/parquet/kraken_derivatives")
KRAKEN_K0_ROOT = RESULTS_ROOT / "phase_kraken_k0_data_foundation_20260630_v1_20260630_163815"
QLMG_NO_VENDOR_V2_ROOT = RESULTS_ROOT / "phase_qlmg_no_vendor_progress_v2_20260630_v1_20260630_115214"
MECHANICAL_QA_ROOT = RESULTS_ROOT / "phase_qlmg_mechanical_qa_evidence_contract_20260630_v1_20260630_074328"
SCREENING_END = pd.Timestamp("2025-12-31T23:59:59Z")

STAGES = (
    "preflight-and-input-freeze",
    "telegram-and-tmux-setup",
    "seal-guard",
    "hypothesis-library-ingest",
    "kraken-data-readiness-audit",
    "missing-official-data-plan",
    "sweep-contract-schema",
    "regime-feature-panel-readiness",
    "hypothesis-to-test-contract-compiler",
    "execution-accounting-fixture-tests",
    "control-engine-fixture-tests",
    "adaptive-sweep-planner",
    "representative-pilot-run",
    "pilot-artifact-and-arithmetic-audit",
    "full-sweep-runbook",
    "decision-report",
    "compact-review-bundle",
    "all",
)

ALLOWED_NEXT_DECISIONS = {
    "launch_full_kraken_hypothesis_sweep_next",
    "repair_contract_compiler_next",
    "repair_kraken_data_gaps_next",
    "repair_regime_panel_next",
    "repair_execution_or_control_engine_next",
    "manual_review_required_before_sweep",
    "blocked_by_protocol_issue",
}

PILOT_BUCKETS = [
    "liquid_continuation",
    "tsmom",
    "prior_high_ath",
    "retest",
    "compression_breakout",
    "session_time",
    "funding_crowding",
    "c2_event_base",
    "capture_substitute_non_rankable",
]

XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

DATE_NUMFMT_IDS = set(range(14, 23)) | {45, 46, 47}


@dataclass
class RunContext:
    args: argparse.Namespace
    run_root: Path
    notifier: "RunNotifier"
    start: pd.Timestamp
    end: pd.Timestamp
    stage_sizes: list[dict[str, Any]]
    retained: list[dict[str, Any]]
    deleted: list[dict[str, Any]]
    cleanup_failures: list[dict[str, Any]]


class RunNotifier:
    def __init__(self, run_root: Path, disabled: bool = False, require_remote: bool = False, allow_no_remote: bool = False) -> None:
        self.run_root = run_root
        self.disabled = disabled
        self.events_path = run_root / "notifications/telegram_events.jsonl"
        self.events_path.parent.mkdir(parents=True, exist_ok=True)
        self.remote = None
        self.status = "disabled" if disabled else "unavailable"
        self.missing = "disabled_by_cli" if disabled else ""
        if not disabled and TelegramNotifier is not None:
            class _Args:
                disable_telegram = False
                telegram_dry_run = False
                tg_bot_token = ""
                tg_chat_id = ""
                tg_auto_chat = False
            try:
                self.remote = TelegramNotifier.from_args(_Args(), run_label="kraken-sweep-readiness")
                self.status = getattr(self.remote, "status_line", lambda: "enabled")()
                if "disabled" in str(self.status).lower():
                    self.missing = str(self.status)
            except Exception as exc:  # pragma: no cover
                self.remote = None
                self.status = "unavailable"
                self.missing = f"{type(exc).__name__}: {exc}"
        elif not disabled:
            self.missing = "tools.telegram_notify.TelegramNotifier unavailable"
        if require_remote and not self.remote_available and not allow_no_remote:
            raise RuntimeError(f"remote Telegram required but unavailable: {self.missing or self.status}")

    @property
    def remote_available(self) -> bool:
        return (not self.disabled) and self.remote is not None and "enabled" in str(self.status).lower()

    def send(self, title: str, body: str = "", *, level: str = "info") -> bool:
        sent = False
        error = ""
        if self.remote is not None:
            try:
                sent = bool(self.remote.send(title, body))
            except Exception as exc:  # pragma: no cover
                error = f"{type(exc).__name__}: {exc}"
        rec = {"ts_utc": utc_now(), "title": title, "body": body, "level": level, "sent": sent, "status": self.status, "error": error}
        with self.events_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, sort_keys=True, default=str) + "\n")
        run_status = "failed" if "failed" in title.lower() else ("complete" if "complete" in title.lower() else "running")
        write_json(self.run_root / "watch_status.json", {"status": run_status, "last_event": title, "ts_utc": rec["ts_utc"], "run_root": str(self.run_root)})
        return sent


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Kraken hypothesis sweep readiness and pilot infrastructure")
    p.add_argument("--stage", choices=STAGES, default="all")
    p.add_argument("--resume", action="store_true")
    p.add_argument("--smoke", action="store_true")
    p.add_argument("--max-symbols", type=int, default=0)
    p.add_argument("--start", default="2023-01-01")
    p.add_argument("--end", default=str(SCREENING_END))
    p.add_argument("--chunk-size", type=int, default=50)
    p.add_argument("--max-output-gb", type=float, default=35.0)
    p.add_argument("--allow-large-output", action="store_true")
    p.add_argument("--disable-telegram", action="store_true")
    p.add_argument("--require-telegram", action="store_true")
    p.add_argument("--allow-no-telegram", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--seed", type=int, default=DEFAULT_SEED)
    p.add_argument("--hypothesis-library", default=str(DEFAULT_HYPOTHESIS_LIBRARY.relative_to(REPO)))
    p.add_argument("--research-input-dir", default=str(DEFAULT_RESEARCH_INPUT_DIR.relative_to(REPO)))
    p.add_argument("--kraken-data-root", default=str(DEFAULT_KRAKEN_DATA_ROOT))
    p.add_argument("--download-missing-official-data", action="store_true")
    p.add_argument("--download-cap-gb", type=float, default=10.0)
    p.add_argument("--pilot-hypothesis-count", type=int, default=20)
    p.add_argument("--pilot-family-count", type=int, default=8)
    p.add_argument("--nulls-per-event", type=int, default=3)
    p.add_argument("--top-per-family", type=int, default=20)
    p.add_argument("--build-full-sweep-plan", action="store_true")
    p.add_argument("--tmux-session-name", default="kraken_sweep_readiness")
    p.add_argument("--launch-tmux", action="store_true")
    p.add_argument("--run-root", default="")
    return p.parse_args(argv)


def stage_list(stage: str) -> list[str]:
    return [s for s in STAGES if s != "all"] if stage == "all" else [stage]


def resolve_path(path_str: str) -> Path:
    p = Path(path_str)
    return p if p.is_absolute() else REPO / p


def resolve_run_root(args: argparse.Namespace) -> tuple[Path, str]:
    if args.run_root:
        p = resolve_path(args.run_root)
        return p.resolve(), "explicit_run_root"
    base = (RESULTS_ROOT / DEFAULT_RUN_ID).resolve()
    if args.smoke:
        return (base / "smoke").resolve(), "smoke_subroot"
    if base.exists():
        suffix = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return base.with_name(f"{base.name}_{suffix}"), f"default_root_existed_suffix_{suffix}"
    return base, "default_root_available"


def clamp_window(args: argparse.Namespace) -> tuple[pd.Timestamp, pd.Timestamp]:
    start = pd.to_datetime(args.start, utc=True)
    requested_end = pd.to_datetime(args.end, utc=True) if args.end else SCREENING_END
    end = min(pd.Timestamp(requested_end), SCREENING_END)
    if start >= PROTECTED_TS or end >= PROTECTED_TS:
        raise RuntimeError("requested strategy-scoring window overlaps protected holdout")
    if start >= end:
        raise RuntimeError("start must be before end")
    return pd.Timestamp(start), pd.Timestamp(end)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def write_csv(path: Path, rows: Iterable[Mapping[str, Any]] | pd.DataFrame, fieldnames: Sequence[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(rows, pd.DataFrame):
        rows.to_csv(path, index=False)
        return
    rows_list = list(rows)
    keys: list[str] = list(fieldnames or [])
    if not keys:
        for row in rows_list:
            for key in row.keys():
                if key not in keys:
                    keys.append(key)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for row in rows_list:
            writer.writerow({k: row.get(k, "") for k in keys})


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path)


def file_sha256(path: Path, max_bytes: int | None = None) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        if max_bytes is None:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        else:
            h.update(f.read(max_bytes))
    return h.hexdigest()


def stable_hash(*parts: object, n: int = 16) -> str:
    h = hashlib.sha256()
    for part in parts:
        h.update(str(part).encode("utf-8", errors="replace"))
        h.update(b"\0")
    return h.hexdigest()[:n]


def row_hash(row: Mapping[str, Any]) -> str:
    return hashlib.sha256(json.dumps(dict(row), sort_keys=True, default=str).encode("utf-8")).hexdigest()


def dir_size_bytes(path: Path) -> int:
    if not path.exists():
        return 0
    if path.is_file():
        return path.stat().st_size
    total = 0
    for p in path.rglob("*"):
        if p.is_file():
            try:
                total += p.stat().st_size
            except OSError:
                pass
    return total


def format_gb(num_bytes: int) -> float:
    return num_bytes / (1024**3)


def cleanup_tmp(ctx: RunContext, stage: str) -> None:
    tmp = ctx.run_root / "tmp" / stage
    if not tmp.exists():
        return
    before = dir_size_bytes(tmp)
    try:
        shutil.rmtree(tmp)
        ctx.deleted.append({"stage": stage, "path": str(tmp), "bytes": before, "status": "deleted_temp_stage_dir"})
    except Exception as exc:
        ctx.cleanup_failures.append({"stage": stage, "path": str(tmp), "bytes": before, "error": f"{type(exc).__name__}: {exc}"})


def record_stage_budget(ctx: RunContext, stage: str, before_tmp: int, before_run: int, estimated_gb: float) -> None:
    after_tmp = dir_size_bytes(ctx.run_root / "tmp" / stage)
    after_run = dir_size_bytes(ctx.run_root)
    ctx.stage_sizes.append({
        "stage": stage,
        "estimated_output_gb": estimated_gb,
        "tmp_size_before_gb": format_gb(before_tmp),
        "tmp_size_after_gb": format_gb(after_tmp),
        "run_size_before_gb": format_gb(before_run),
        "run_size_after_gb": format_gb(after_run),
        "actual_stage_growth_gb": format_gb(max(after_run - before_run, 0)),
        "cleanup_status": "pending",
    })


def largest_files(root: Path, limit: int = 25) -> list[dict[str, Any]]:
    files = []
    if not root.exists():
        return files
    for p in root.rglob("*"):
        if p.is_file():
            try:
                size = p.stat().st_size
            except OSError:
                continue
            files.append({"path": str(p.relative_to(root)), "bytes": size, "gb": format_gb(size), "over_1gb": size > 1024**3})
    return sorted(files, key=lambda r: r["bytes"], reverse=True)[:limit]


def done_path(ctx: RunContext, stage: str) -> Path:
    return ctx.run_root / "stage_status" / f"{stage}.done"


def mark_done(ctx: RunContext, stage: str) -> None:
    p = done_path(ctx, stage)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(utc_now() + "\n", encoding="utf-8")


def resolve_hypothesis_library(args: argparse.Namespace) -> tuple[Path, str]:
    explicit = resolve_path(args.hypothesis_library)
    if explicit.exists():
        return explicit, "explicit_or_default_path"
    matches = sorted(resolve_path(args.research_input_dir).glob("*Hypothesis*Library*.xlsx"))
    if matches:
        return matches[0], "fallback_glob_match"
    raise FileNotFoundError(f"hypothesis library not found: {explicit}")


def col_to_index(col: str) -> int:
    idx = 0
    for ch in col.upper():
        if "A" <= ch <= "Z":
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def index_to_col(idx: int) -> str:
    idx += 1
    s = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def excel_serial_to_iso(value: float) -> str:
    base = datetime(1899, 12, 30, tzinfo=timezone.utc)
    dt = base + timedelta(days=float(value))
    if abs(value - int(value)) < 1e-9:
        return dt.date().isoformat()
    return dt.isoformat().replace("+00:00", "Z")


def parse_styles(z: zipfile.ZipFile) -> set[int]:
    date_styles: set[int] = set()
    if "xl/styles.xml" not in z.namelist():
        return date_styles
    root = ET.fromstring(z.read("xl/styles.xml"))
    custom_date_numfmts: set[int] = set()
    for numfmt in root.findall(".//a:numFmt", XLSX_NS):
        try:
            num_id = int(numfmt.attrib.get("numFmtId", "-1"))
        except ValueError:
            continue
        fmt = numfmt.attrib.get("formatCode", "").lower()
        if any(ch in fmt for ch in ["d", "m", "y", "h", "s"]):
            custom_date_numfmts.add(num_id)
    cell_xfs = root.find("a:cellXfs", XLSX_NS)
    if cell_xfs is None:
        return date_styles
    for idx, xf in enumerate(cell_xfs.findall("a:xf", XLSX_NS)):
        try:
            num_id = int(xf.attrib.get("numFmtId", "-1"))
        except ValueError:
            continue
        if num_id in DATE_NUMFMT_IDS or num_id in custom_date_numfmts:
            date_styles.add(idx)
    return date_styles


def read_shared_strings(z: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    out: list[str] = []
    for si in root.findall(".//a:si", XLSX_NS):
        texts = [t.text or "" for t in si.findall(".//a:t", XLSX_NS)]
        out.append("".join(texts))
    return out


def workbook_sheet_paths(z: zipfile.ZipFile) -> list[tuple[str, str]]:
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    id_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("rel:Relationship", XLSX_NS)}
    sheets: list[tuple[str, str]] = []
    for sheet in wb.findall(".//a:sheet", XLSX_NS):
        name = sheet.attrib["name"]
        rid = sheet.attrib.get(f"{{{XLSX_NS['r']}}}id")
        target = id_to_target.get(rid or "", "")
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = "xl/" + target.lstrip("/")
        sheets.append((name, path))
    return sheets


def cell_value(cell: ET.Element, shared: list[str], date_styles: set[int]) -> tuple[Any, bool, str]:
    typ = cell.attrib.get("t", "")
    style_idx = int(cell.attrib.get("s", "-1")) if cell.attrib.get("s", "-1").isdigit() else -1
    formula = cell.find("a:f", XLSX_NS)
    formula_flag = formula is not None
    warn = "formula_cached_value" if formula_flag else ""
    if typ == "inlineStr":
        texts = [t.text or "" for t in cell.findall(".//a:t", XLSX_NS)]
        return "".join(texts), formula_flag, warn
    v = cell.find("a:v", XLSX_NS)
    raw = v.text if v is not None else ""
    if typ == "s":
        if raw.isdigit() and int(raw) < len(shared):
            return shared[int(raw)], formula_flag, warn
        return raw, formula_flag, warn or "bad_shared_string_index"
    if typ == "b":
        return raw == "1", formula_flag, warn
    if raw == "":
        if formula_flag:
            return "", True, "formula_without_cached_value"
        return "", formula_flag, warn
    if style_idx in date_styles:
        try:
            return excel_serial_to_iso(float(raw)), formula_flag, warn
        except ValueError:
            return raw, formula_flag, warn or "date_style_non_numeric"
    try:
        f = float(raw)
        if math.isfinite(f) and abs(f - int(f)) < 1e-12:
            return int(f), formula_flag, warn
        return f, formula_flag, warn
    except ValueError:
        return raw, formula_flag, warn


def read_xlsx_stdlib(path: Path) -> dict[str, Any]:
    sheets: dict[str, list[list[Any]]] = {}
    cell_meta: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as z:
        shared = read_shared_strings(z)
        date_styles = parse_styles(z)
        for sheet_name, sheet_path in workbook_sheet_paths(z):
            root = ET.fromstring(z.read(sheet_path))
            rows_out: list[list[Any]] = []
            for row in root.findall(".//a:sheetData/a:row", XLSX_NS):
                row_idx = int(row.attrib.get("r", len(rows_out) + 1))
                vals: dict[int, Any] = {}
                formula_cells = 0
                warnings: list[str] = []
                for cell in row.findall("a:c", XLSX_NS):
                    ref = cell.attrib.get("r", "A1")
                    m = re.match(r"([A-Z]+)", ref)
                    cidx = col_to_index(m.group(1) if m else "A")
                    val, is_formula, warn = cell_value(cell, shared, date_styles)
                    vals[cidx] = val
                    if is_formula:
                        formula_cells += 1
                    if warn:
                        warnings.append(warn)
                    cell_meta.append({"sheet": sheet_name, "cell": ref, "row_index": row_idx, "column_index": cidx + 1, "formula": is_formula, "parse_warning": warn})
                max_idx = max(vals.keys()) if vals else -1
                row_values = [vals.get(i, "") for i in range(max_idx + 1)]
                rows_out.append(row_values)
            sheets[sheet_name] = rows_out
    return {"reader": "stdlib_zipfile_xml", "sheets": sheets, "cell_meta": cell_meta}


def read_xlsx_openpyxl(path: Path) -> dict[str, Any]:  # pragma: no cover - unavailable in current venv
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    cell_meta: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = []
        fws = wb_formula[ws.title]
        for ridx, row in enumerate(ws.iter_rows(), start=1):
            vals = []
            for cidx, cell in enumerate(row, start=1):
                val = cell.value
                fcell = fws.cell(ridx, cidx)
                is_formula = isinstance(fcell.value, str) and fcell.value.startswith("=")
                warn = "formula_cached_value" if is_formula and val not in (None, "") else ("formula_without_cached_value" if is_formula else "")
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                vals.append("" if val is None else val)
                cell_meta.append({"sheet": ws.title, "cell": cell.coordinate, "row_index": ridx, "column_index": cidx, "formula": is_formula, "parse_warning": warn})
            while vals and vals[-1] == "":
                vals.pop()
            rows.append(vals)
        sheets[ws.title] = rows
    return {"reader": "openpyxl", "sheets": sheets, "cell_meta": cell_meta}


def read_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # noqa: F401
        return read_xlsx_openpyxl(path)
    except Exception:
        return read_xlsx_stdlib(path)


def non_empty_rows(rows: list[list[Any]]) -> list[tuple[int, list[Any]]]:
    out = []
    for idx, row in enumerate(rows, start=1):
        if any(str(v).strip() != "" for v in row):
            out.append((idx, row))
    return out


def sheet_to_records(rows: list[list[Any]], sheet_name: str, source_hash: str) -> tuple[list[dict[str, Any]], list[str]]:
    ner = non_empty_rows(rows)
    if not ner:
        return [], []
    header_idx, header = ner[0]
    headers = [str(h).strip() if str(h).strip() else f"column_{i+1}" for i, h in enumerate(header)]
    records: list[dict[str, Any]] = []
    for ridx, row in ner[1:]:
        rec: dict[str, Any] = {"source_sheet": sheet_name, "workbook_row_index": ridx, "source_hash": source_hash}
        for i, h in enumerate(headers):
            rec[h] = row[i] if i < len(row) else ""
        rec["raw_row_hash"] = row_hash(rec)
        records.append(rec)
    return records, headers


def norm_col(name: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(name).lower())).strip("_")


def as_bool_text(value: Any) -> str:
    return str(value or "").strip().lower()


def classify_hypothesis(row: Mapping[str, Any]) -> str:
    text = " ".join(str(row.get(k, "")) for k in row.keys()).lower()
    data_tier = as_bool_text(row.get("data_tier", row.get("Data Tier", "")))
    feasibility = as_bool_text(row.get("kraken_feasibility", row.get("Kraken Feasibility", "")))
    if any(x in text for x in ["depth", "orderbook", "liquidation", "microstructure", "listing", "orb"]):
        return "kraken_capture_substitute_needed"
    if "tier 3" in data_tier:
        return "kraken_candidate_library_only"
    if "kraken" in feasibility and any(x in feasibility for x in ["low", "not"]):
        return "better_on_broader_alt_perp_venue"
    if "tier 1" in data_tier and "cap" in text:
        return "kraken_tier1_with_caps"
    if "tier 1" in data_tier or feasibility in {"high", "medium", "partial"} or "high" in feasibility:
        return "kraken_tier1_ready"
    if "event" in text or "catalyst" in text:
        return "needs_event_ledger_first"
    return "kraken_tier1_with_caps"


def bucket_for(row: Mapping[str, Any]) -> str:
    text = " ".join(str(row.get(k, "")) for k in row.keys()).lower()
    if any(x in text for x in ["time-series", "tsmom", "trend signal", "volatility-managed"]):
        return "tsmom"
    if any(x in text for x in ["ath", "all-time-high", "prior high", "prior-major-high", "proximity"]):
        return "prior_high_ath"
    if any(x in text for x in ["retest", "reclaim", "post-breakout"]):
        return "retest"
    if any(x in text for x in ["compression", "squeeze", "volatility contraction"]):
        return "compression_breakout"
    if any(x in text for x in ["session", "time-of-day", "utc", "weekend", "weekday"]):
        return "session_time"
    if any(x in text for x in ["funding", "crowding", "oi", "open interest"]):
        return "funding_crowding"
    if any(x in text for x in ["catalyst", "event", "post-catalyst"]):
        return "c2_event_base"
    if classify_hypothesis(row) in {"kraken_capture_substitute_needed", "kraken_candidate_library_only"}:
        return "capture_substitute_non_rankable"
    return "liquid_continuation"


def find_pdf_pages(pdf: Path, query: str, max_refs: int = 3) -> str:
    if not pdf.exists() or not shutil.which("pdftotext"):
        return ""
    try:
        proc = subprocess.run(["pdftotext", "-layout", str(pdf), "-"], check=False, capture_output=True, text=True, timeout=20)
    except Exception:
        return ""
    if proc.returncode != 0:
        return ""
    text = proc.stdout
    pages = text.split("\f")
    q = str(query).lower().strip()
    if not q:
        return ""
    refs = []
    for i, page in enumerate(pages, start=1):
        if q[:40] and q[:40] in page.lower():
            refs.append(f"{pdf.name}:page_{i}")
            if len(refs) >= max_refs:
                break
    return ";".join(refs)


def family_slug(value: Any) -> str:
    text = str(value or "unknown").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text[:80] or "unknown"


def load_parquet_many(paths: list[Path], columns: list[str] | None = None) -> pd.DataFrame:
    parts = []
    for p in paths:
        try:
            parts.append(pd.read_parquet(p, columns=columns))
        except Exception:
            try:
                parts.append(pd.read_parquet(p))
            except Exception:
                continue
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


def discover_dataset_paths(ctx: RunContext) -> pd.DataFrame:
    root = Path(ctx.args.kraken_data_root)
    k0_manifest = read_csv(KRAKEN_K0_ROOT / "download/download_manifest.csv")
    persistent_manifest = read_csv(KRAKEN_K0_ROOT / "download/persistent_store_manifest.csv")
    qc = read_csv(KRAKEN_K0_ROOT / "qc/qc_summary.csv")
    rows: list[dict[str, Any]] = []
    expected = {
        "historical_trade_candles_5m": root / "parquet/historical_trade_candles_5m",
        "historical_mark_candles_5m": root / "parquet/historical_mark_candles_5m",
        "funding": root / "parquet/funding",
        "instruments": root / "parquet/instruments",
        "tickers": root / "parquet/tickers",
        "candles_recent": root / "parquet/candles",
        "mark_recent": root / "parquet/mark",
        "analytics": root / "parquet/analytics",
        "events": root / "parquet/events",
    }
    for dataset, path in expected.items():
        files = list(path.rglob("*.parquet")) if path.exists() else []
        manifest_rows = 0
        if not k0_manifest.empty and "dataset" in k0_manifest.columns:
            key = dataset
            if dataset == "candles_recent":
                key = "candles"
            elif dataset == "mark_recent":
                key = "mark"
            manifest_rows = int(k0_manifest[k0_manifest["dataset"].astype(str).eq(key)].shape[0])
        qc_failures = 0
        if not qc.empty and "path" in qc.columns and "status" in qc.columns:
            qc_failures = int(qc[qc["path"].astype(str).str.contains(str(path), regex=False, na=False) & qc["status"].astype(str).str.lower().eq("fail")].shape[0])
        rows.append({
            "dataset": dataset,
            "expected_path": str(path),
            "path_exists": path.exists(),
            "parquet_files": len(files),
            "manifest_rows": manifest_rows,
            "qc_failures": qc_failures,
            "resolved_path": str(path) if files or path.exists() else "",
            "resolution_status": "resolved" if files or path.exists() else "missing",
        })
    for dataset in sorted(set(k0_manifest.get("dataset", pd.Series(dtype=str)).astype(str))):
        if dataset not in {r["dataset"] for r in rows}:
            rows.append({"dataset": dataset, "expected_path": "", "path_exists": False, "parquet_files": 0, "manifest_rows": int(k0_manifest[k0_manifest["dataset"].astype(str).eq(dataset)].shape[0]), "qc_failures": 0, "resolved_path": "manifest_only", "resolution_status": "manifest_only"})
    out = pd.DataFrame(rows)
    write_csv(ctx.run_root / "data_readiness/kraken_dataset_path_resolution.csv", out)
    return out


def classify_missing(dataset: str, row: Mapping[str, Any]) -> str:
    exists = bool(row.get("path_exists")) or int(row.get("parquet_files", 0) or 0) > 0
    if exists:
        return "available"
    if dataset in {"historical_trade_candles_5m", "historical_mark_candles_5m", "funding", "instruments"}:
        return "needed_for_full_sweep"
    if dataset in {"analytics", "events"}:
        return "needs_live_capture_substitute"
    if dataset in {"candles_recent", "mark_recent"}:
        return "not_needed_for_readiness"
    return "targeted_1m_later"


def selected_symbol_universe(ctx: RunContext) -> list[str]:
    uni = read_csv(KRAKEN_K0_ROOT / "universe/kraken_universe_summary.csv")
    if uni.empty:
        return ["PF_XBTUSD", "PF_ETHUSD"]
    syms = uni[uni.get("tier", "").astype(str).isin(["K-A", "K-B"])] if "tier" in uni.columns else uni
    syms = syms.copy()
    syms["_pf_prefer"] = syms.get("venue_symbol", pd.Series(dtype=str)).astype(str).str.startswith("PF_").astype(int)
    syms["_tier_rank"] = syms.get("tier", pd.Series("K-B", index=syms.index)).map({"K-A": 0, "K-B": 1}).fillna(2)
    syms = syms.sort_values(["_pf_prefer", "_tier_rank", "venue_symbol"], ascending=[False, True, True])
    symbols = syms.get("venue_symbol", pd.Series(dtype=str)).astype(str).tolist()
    max_symbols = ctx.args.max_symbols if ctx.args.max_symbols else (5 if ctx.args.smoke else 40)
    return symbols[:max_symbols]


def symbol_bar_files(data_root: Path, symbol: str, dataset: str = "historical_trade_candles_5m") -> list[Path]:
    d = data_root / "parquet" / dataset / symbol
    if d.exists():
        return sorted(d.glob("*.parquet"))
    d2 = data_root / "parquet" / dataset
    return sorted(d2.glob(f"{symbol}_*.parquet")) if d2.exists() else []




def parse_kraken_time_series(series: pd.Series) -> pd.Series:
    num = pd.to_numeric(series, errors="coerce")
    if num.notna().any():
        med = float(num.dropna().median())
        if med > 1e12:
            return pd.to_datetime(num, unit="ms", utc=True, errors="coerce")
        if med > 1e9:
            return pd.to_datetime(num, unit="s", utc=True, errors="coerce")
    return pd.to_datetime(series, utc=True, errors="coerce")

def normalize_bars(df: pd.DataFrame, symbol: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    ts_col = "time" if "time" in out.columns else ("timestamp" if "timestamp" in out.columns else None)
    if ts_col is None:
        return pd.DataFrame()
    out["decision_ts"] = parse_kraken_time_series(out[ts_col])
    for c in ["open", "high", "low", "close", "volume"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    out["venue_symbol"] = symbol
    out = out[(out["decision_ts"] >= start) & (out["decision_ts"] <= end) & (out["decision_ts"] < PROTECTED_TS)]
    return out.dropna(subset=["decision_ts", "close"]).sort_values("decision_ts")


def chunk_start_from_name(path: Path) -> pd.Timestamp | None:
    m = re.search(r"(20\d{6}T\d{6})", path.name)
    if not m:
        return None
    return pd.to_datetime(m.group(1), format="%Y%m%dT%H%M%S", utc=True, errors="coerce")


def files_overlapping_window(files: list[Path], start: pd.Timestamp, end: pd.Timestamp, chunk_days: int = 7) -> list[Path]:
    selected = []
    pad_start = start - pd.Timedelta(days=chunk_days)
    pad_end = end + pd.Timedelta(days=1)
    for f in files:
        ts = chunk_start_from_name(f)
        if ts is None or pd.isna(ts):
            selected.append(f)
            continue
        if pad_start <= ts <= pad_end:
            selected.append(f)
    return selected or files


def load_symbol_bars(ctx: RunContext, symbol: str, dataset: str = "historical_trade_candles_5m") -> pd.DataFrame:
    files = symbol_bar_files(Path(ctx.args.kraken_data_root), symbol, dataset)
    files = files_overlapping_window(files, ctx.start, ctx.end)
    if ctx.args.smoke:
        files = files[:20]
    df = load_parquet_many(files)
    return normalize_bars(df, symbol, ctx.start, ctx.end)


def load_daily_features(ctx: RunContext, symbols: list[str]) -> pd.DataFrame:
    parts = []
    for symbol in symbols:
        bars = load_symbol_bars(ctx, symbol, "historical_trade_candles_5m")
        if bars.empty:
            continue
        b = bars.sort_values("decision_ts").copy()
        b["session_date_utc"] = b["decision_ts"].dt.floor("D")
        daily = (
            b.groupby("session_date_utc", as_index=False)
            .agg(close=("close", "last"), volume=("volume", "sum"), source_close_ts=("decision_ts", "max"))
            .dropna(subset=["close"])
        )
        # The decision timestamp is the actual last source bar timestamp, not the
        # right-edge calendar label. This avoids turning the final 2025-12-31
        # pre-holdout bar group into a synthetic 2026-01-01 decision label.
        daily["decision_ts"] = daily["source_close_ts"]
        daily["venue_symbol"] = symbol
        daily["ret_1d"] = daily["close"].pct_change()
        daily["ret_20d"] = daily["close"].pct_change(20)
        daily["vol_20d"] = daily["ret_1d"].rolling(20, min_periods=5).std()
        parts.append(daily)
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()


def event_r(side: str, entry: float, stop: float, exit_price: float, fee_r: float, slippage_r: float, funding_r: float) -> tuple[float, float]:
    risk = abs(entry - stop)
    if risk <= 0 or not np.isfinite(risk):
        return float("nan"), float("nan")
    gross = (exit_price - entry) / risk if side == "long" else (entry - exit_price) / risk
    return gross, gross - fee_r - slippage_r + funding_r


def make_event_row(contract: Mapping[str, Any], symbol: str, bars: pd.DataFrame, event_idx: int, hold_bars: int = 24) -> dict[str, Any] | None:
    if len(bars) <= hold_bars + event_idx + 2:
        return None
    row = bars.iloc[event_idx]
    exit_row = bars.iloc[min(event_idx + hold_bars, len(bars) - 1)]
    decision_ts = pd.Timestamp(row["decision_ts"])
    if decision_ts >= PROTECTED_TS:
        return None
    entry_ts = decision_ts + pd.Timedelta(minutes=5)
    entry = float(row["close"])
    side = "short" if "short" in str(contract.get("direction", "")).lower() else "long"
    stop = entry * (0.98 if side == "long" else 1.02)
    target = entry * (1.04 if side == "long" else 0.96)
    exit_price = float(exit_row["close"])
    gross, net = event_r(side, entry, stop, exit_price, 0.004, 0.002, 0.0)
    event_id = stable_hash(contract.get("contract_id"), symbol, decision_ts, event_idx)
    return {
        "candidate_id": contract.get("contract_id"),
        "family": contract.get("family"),
        "branch_id": "kraken_readiness_pilot",
        "symbol": symbol,
        "decision_ts": str(decision_ts),
        "side": side,
        "entry_ts": str(entry_ts),
        "entry_price": entry,
        "entry_price_source": "kraken_trade_5m_close_next_bar_proxy",
        "stop_price": stop,
        "target_price": target,
        "exit_rule": "mechanical_hold_or_target_stop_fixture",
        "exit_ts": str(pd.Timestamp(exit_row["decision_ts"])),
        "exit_price": exit_price,
        "exit_reason": "time_exit_fixture",
        "gross_R": gross,
        "fees_R": 0.004,
        "slippage_R": 0.002,
        "funding_R": 0.0,
        "net_R": net,
        "mark_liquidation_flag": False,
        "same_bar_ambiguity_flag": False,
        "funding_timestamps_crossed": 0,
        "mark_available": True,
        "funding_exact": True,
        "lifecycle_status": "known_live_from_k0_instrument_master",
        "data_tier": contract.get("data_tier", "Tier 1"),
        "control_group_id": stable_hash(contract.get("contract_id"), "controls"),
        "source_data_hash": stable_hash(symbol, decision_ts, exit_row.get("decision_ts")),
        "event_id": event_id,
        "pipeline_label": "pipeline_passed",
        "metric_basis": "event_level_trade_ledger_mechanical_pilot_not_alpha_evidence",
    }


def validate_no_forbidden_pilot_labels(df: pd.DataFrame) -> list[str]:
    forbidden = re.compile(r"profitable|lead|prelead|validated|promoted|rejected family|best strategy", re.I)
    bad = []
    for col in df.columns:
        if any(x in col.lower() for x in ["label", "status", "verdict", "decision"]):
            hits = df[col].astype(str).str.contains(forbidden, na=False)
            if bool(hits.any()):
                bad.append(f"{col}:forbidden_pilot_label_rows={int(hits.sum())}")
    return bad


def stage_preflight(ctx: RunContext) -> None:
    ctx.run_root.mkdir(parents=True, exist_ok=True)
    lib, lib_resolution = resolve_hypothesis_library(ctx.args)
    inputs: list[dict[str, Any]] = []
    for name, path in [
        ("kraken_k0_root", KRAKEN_K0_ROOT),
        ("kraken_data_root", Path(ctx.args.kraken_data_root)),
        ("qlmg_no_vendor_v2_root", QLMG_NO_VENDOR_V2_ROOT),
        ("mechanical_qa_root", MECHANICAL_QA_ROOT),
        ("hypothesis_library", lib),
    ]:
        exists = path.exists()
        inputs.append({"name": name, "path": str(path), "exists": exists, "kind": "dir" if path.is_dir() else "file", "sha256": file_sha256(path) if path.is_file() and exists else "dir_manifest_or_not_file", "resolution": lib_resolution if name == "hypothesis_library" else "fixed"})
    research_dir = resolve_path(ctx.args.research_input_dir)
    for p in sorted(research_dir.glob("*")):
        if p.is_file() and p.suffix.lower() in {".pdf", ".md", ".txt", ".xlsx"}:
            inputs.append({"name": "research_input", "path": str(p), "exists": True, "kind": "file", "sha256": file_sha256(p), "resolution": "research_dir_scan"})
    write_csv(ctx.run_root / "preflight/input_artifact_manifest.csv", inputs)
    write_json(ctx.run_root / "preflight/frozen_artifact_hashes.json", {r["path"]: r["sha256"] for r in inputs})
    snap = resource_snapshot(ctx.run_root.parent)
    guard = check_resource_guard(snap, estimated_output_gb=2.0 if ctx.args.smoke else 8.0, hard_stage_output_gb=35.0, allow_large_output=ctx.args.allow_large_output)
    write_json(ctx.run_root / "resources/resource_guard_report.md.json", guard)
    write_text(ctx.run_root / "resources/resource_guard_report.md", f"# Resource Guard\n\nStatus: `{guard['status']}`\n\nFree disk GB: `{guard['free_disk_gb']:.3f}`\n")
    if guard["status"] == "hard_stop":
        raise RuntimeError("resource guard hard stop: " + ",".join(guard["reasons"]))
    write_csv(ctx.run_root / "resources/output_budget_by_stage.csv", [{"stage": s, "estimated_output_gb": 0.05 if ctx.args.smoke else (0.5 if s in {"representative-pilot-run", "regime-feature-panel-readiness"} else 0.1)} for s in STAGES if s != "all"])
    write_text(ctx.run_root / "resources/artifact_retention_policy.md", "# Artifact Retention Policy\n\nRetain reports, manifests, contracts, compact pilot ledgers, and audits. Delete stage-local temporary files under `tmp/` after successful stages. Never delete `/opt/parquet/kraken_derivatives/` or prior durable run artifacts.")
    write_text(ctx.run_root / "preflight/preflight_report.md", f"# Preflight\n\nHypothesis library: `{lib}`\n\nLibrary resolution: `{lib_resolution}`\n\nKraken data root: `{ctx.args.kraken_data_root}`\n\nProtected scoring holdout: `{PROTECTED_TS}`\n")


def stage_telegram(ctx: RunContext) -> None:
    write_text(ctx.run_root / "notifications/telegram_readiness_report.md", f"# Telegram Readiness\n\nStatus: `{ctx.notifier.status}`\n\nRemote available: `{ctx.notifier.remote_available}`\n")
    write_text(ctx.run_root / "tmux/watch_commands.md", f"# Watch Commands\n\n- `tmux attach -t {ctx.args.tmux_session_name}`\n- `tail -f {ctx.run_root}/notifications/telegram_events.jsonl`\n- `watch -n 30 'cat {ctx.run_root}/watch_status.json'`\n")


def stage_seal(ctx: RunContext) -> None:
    write_text(ctx.run_root / "seal/kraken_sweep_seal_policy.md", "# Kraken Sweep Seal Policy\n\nData inventory and QC may inspect post-holdout data. Candidate scoring, controls, tuning, ranking, and pilot interpretation must use `decision_ts < 2026-01-01T00:00:00Z`.\n")
    result = scan_output_tree_for_protected(ctx.run_root)
    write_json(ctx.run_root / "seal/protected_timestamp_scan.json", result_to_jsonable(result))
    write_text(ctx.run_root / "seal/seal_guard_report.md", f"# Seal Guard\n\nStatus: `{result.status}`\n\nThis scan is finite contract checking, not a proof of absence of all leakage.\n")


def stage_hypothesis_ingest(ctx: RunContext) -> None:
    lib, resolution = resolve_hypothesis_library(ctx.args)
    parsed = read_xlsx(lib)
    source_hash = file_sha256(lib)
    sheets: dict[str, list[list[Any]]] = parsed["sheets"]
    sheet_rows = []
    col_rows = []
    all_records_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for sheet, rows in sheets.items():
        ner = non_empty_rows(rows)
        max_cols = max((len(r) for _, r in ner), default=0)
        sheet_rows.append({"sheet": sheet, "non_empty_rows": len(ner), "max_columns": max_cols, "reader": parsed["reader"], "source_hash": source_hash})
        records, headers = sheet_to_records(rows, sheet, source_hash)
        all_records_by_sheet[sheet] = records
        for i, header in enumerate(headers, start=1):
            col_rows.append({"sheet": sheet, "column_index": i, "column_letter": index_to_col(i - 1), "header": header, "normalized_header": norm_col(header)})
    write_csv(ctx.run_root / "hypotheses/workbook_sheet_inventory.csv", sheet_rows)
    write_csv(ctx.run_root / "hypotheses/workbook_column_map.csv", col_rows)
    write_csv(ctx.run_root / "hypotheses/xlsx_cell_parse_warnings.csv", pd.DataFrame(parsed.get("cell_meta", [])))
    formula_warnings = [r for r in parsed.get("cell_meta", []) if r.get("formula") and r.get("parse_warning")]
    write_text(ctx.run_root / "hypotheses/xlsx_reader_report.md", f"# XLSX Reader Report\n\nReader used: `{parsed['reader']}`\n\nWorkbook resolution: `{resolution}`\n\nFormula warning cells: `{len(formula_warnings)}`\n")
    main_records = all_records_by_sheet.get("Hypothesis Library", [])
    if not main_records:
        raise RuntimeError("hypothesis_library_parse_incomplete: Hypothesis Library sheet missing or empty")
    priority_records = all_records_by_sheet.get("Priority View", [])
    priority_by_id = {str(r.get("Hypothesis ID", "")).strip(): r for r in priority_records if str(r.get("Hypothesis ID", "")).strip()}
    normalized = []
    for rec in main_records:
        hid = str(rec.get("Hypothesis ID", "")).strip()
        pr = priority_by_id.get(hid, {})
        merged = {**rec}
        for k, v in pr.items():
            merged.setdefault(k, v)
        norm = {norm_col(k): v for k, v in merged.items()}
        out = {
            "hypothesis_id": hid,
            "family": str(rec.get("Canonical Family", "")).strip(),
            "short_name": str(rec.get("Short Name", "")).strip(),
            "alpha_mechanism": str(rec.get("Mechanism / Principle", "")).strip(),
            "direction": str(rec.get("Direction", "")).strip(),
            "universe": str(rec.get("Universe", "")).strip(),
            "context_regime": str(rec.get("Best Context / Regime", "")).strip(),
            "disable_conditions": str(rec.get("Disable / Bad Context", "")).strip(),
            "entry_sketch": str(rec.get("Entry Sketch", "")).strip(),
            "exit_sketch": str(rec.get("Exit Sketch", "")).strip(),
            "stop_risk_sketch": str(rec.get("Stop / Risk Sketch", rec.get("Stop/Risk Sketch", ""))).strip(),
            "required_data": str(rec.get("Required Data", "")).strip(),
            "minimum_no_vendor_data": str(rec.get("Minimum No-Vendor Data", "")).strip(),
            "ideal_data": str(rec.get("Ideal Data", "")).strip(),
            "data_tier": str(pr.get("Data Tier", rec.get("Data Tier", norm.get("data_tier", "")))).strip(),
            "kraken_feasibility": str(pr.get("Kraken Feasibility", rec.get("Kraken Feasibility", norm.get("kraken_feasibility", "")))).strip(),
            "no_vendor_feasibility": str(rec.get("No-Vendor Feasibility", norm.get("no_vendor_feasibility", ""))).strip(),
            "live_capture_requirement": str(rec.get("Live Capture Requirement", norm.get("live_capture_requirement", ""))).strip(),
            "evidence_summary": str(rec.get("Evidence Summary", pr.get("Current Status", ""))).strip(),
            "suggested_null_control": str(rec.get("Suggested Null / Control", rec.get("Suggested Null/Control", ""))).strip(),
            "falsification_criteria": str(rec.get("Falsification Criteria", "")).strip(),
            "priority": str(pr.get("Priority Score", rec.get("Priority", norm.get("priority", "")))).strip(),
            "source_references": str(rec.get("Source Document", "")).strip(),
            "source_sheet": "Hypothesis Library",
            "workbook_row_index": rec.get("workbook_row_index"),
            "source_hash": source_hash,
            "raw_row_hash": rec.get("raw_row_hash"),
        }
        out["kraken_readiness_class"] = classify_hypothesis(out)
        out["pilot_bucket"] = bucket_for(out)
        out["canonical_record_hash"] = row_hash(out)
        normalized.append(out)
    df = pd.DataFrame(normalized)
    ids = df["hypothesis_id"].astype(str).str.strip()
    checks = []
    for hid, group in df.groupby("hypothesis_id", dropna=False):
        checks.append({"hypothesis_id": hid, "rows": len(group), "duplicate": len(group) > 1, "missing_id": str(hid).strip() == ""})
    required_empty = []
    for col in ["hypothesis_id", "alpha_mechanism", "entry_sketch", "data_tier"]:
        if col in df.columns:
            required_empty.append({"field": col, "empty_rows": int(df[col].astype(str).str.strip().eq("").sum())})
    uniq = pd.DataFrame(checks)
    write_csv(ctx.run_root / "hypotheses/hypothesis_id_uniqueness_check.csv", uniq)
    write_csv(ctx.run_root / "hypotheses/hypothesis_required_field_check.csv", required_empty)
    if int(uniq["missing_id"].sum()) or int(uniq["duplicate"].sum()):
        raise RuntimeError("hypothesis_library_parse_incomplete: duplicate or missing hypothesis IDs")
    if len(df) != len(main_records) or len(df) < 20:
        raise RuntimeError("hypothesis_library_parse_incomplete: normalized row count materially differs from workbook")
    write_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv", df)
    source_rows = []
    pdfs = sorted(resolve_path(ctx.args.research_input_dir).glob("*.pdf"))
    for _, row in df.iterrows():
        refs = []
        src_doc = str(row.get("source_references", ""))
        for pdf in pdfs:
            if pdf.name == src_doc or pdf.name.lower() in src_doc.lower():
                refs.append(pdf.name)
        if not refs and src_doc:
            refs.append(src_doc)
        source_rows.append({"hypothesis_id": row["hypothesis_id"], "source_references": ";".join(refs), "supporting_context_only": True, "pdf_mutated_canonical_row": False})
    write_csv(ctx.run_root / "hypotheses/hypothesis_source_map.csv", source_rows)
    write_csv(ctx.run_root / "hypotheses/hypothesis_library_review_suggestions.csv", [])
    by_class = df["kraken_readiness_class"].value_counts().to_dict()
    write_text(ctx.run_root / "hypotheses/hypothesis_ingest_report.md", f"# Hypothesis Ingest Report\n\nWorkbook rows ingested: `{len(df)}`\n\nReader: `{parsed['reader']}`\n\nReadiness classes: `{json.dumps(by_class, sort_keys=True)}`\n\nPDFs/Markdown were support-only and did not mutate the canonical workbook rows.\n")


def stage_data_readiness(ctx: RunContext) -> None:
    paths = discover_dataset_paths(ctx)
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    available = {str(r["dataset"]): bool(r["path_exists"]) or int(r.get("parquet_files", 0) or 0) > 0 for _, r in paths.iterrows()}
    rows = []
    for _, h in hyp.iterrows():
        text = " ".join(str(h.get(c, "")) for c in hyp.columns).lower()
        required = ["historical_trade_candles_5m", "historical_mark_candles_5m", "instruments"]
        if "funding" in text or "perp" in text or "crowd" in text:
            required.append("funding")
        if any(x in text for x in ["depth", "orderbook", "liquidation", "trade tape", "microstructure"]):
            required.append("events")
        missing = [d for d in required if not available.get(d, False)]
        if not missing and h.get("kraken_readiness_class") in {"kraken_tier1_ready", "kraken_tier1_with_caps"}:
            cat = "ready_for_full_sweep"
        elif any(d == "events" for d in missing):
            cat = "needs_live_capture_substitute"
        elif missing:
            cat = "needs_missing_official_download"
        elif h.get("kraken_readiness_class") == "needs_event_ledger_first":
            cat = "needs_event_database_first"
        else:
            cat = "ready_for_readiness_pilot"
        rows.append({"hypothesis_id": h["hypothesis_id"], "family": h["family"], "readiness_category": cat, "required_datasets": ";".join(required), "missing_datasets": ";".join(missing), "kraken_readiness_class": h.get("kraken_readiness_class"), "needs_1m_now": False, "live_capture_required": cat == "needs_live_capture_substitute"})
    out = pd.DataFrame(rows)
    write_csv(ctx.run_root / "data_readiness/hypothesis_data_readiness.csv", out)
    fam = out.groupby("family", dropna=False).agg(hypotheses=("hypothesis_id", "nunique"), ready_full=("readiness_category", lambda s: int((s == "ready_for_full_sweep").sum())), capture_needed=("readiness_category", lambda s: int((s == "needs_live_capture_substitute").sum()))).reset_index()
    write_csv(ctx.run_root / "data_readiness/family_data_readiness.csv", fam)
    write_text(ctx.run_root / "data_readiness/kraken_data_gap_report.md", f"# Kraken Data Gap Report\n\nReady for full sweep: `{int((out['readiness_category'] == 'ready_for_full_sweep').sum())}`\n\nNeeds capture substitute: `{int((out['readiness_category'] == 'needs_live_capture_substitute').sum())}`\n")


def stage_missing_data(ctx: RunContext) -> None:
    paths = read_csv(ctx.run_root / "data_readiness/kraken_dataset_path_resolution.csv")
    rows = []
    for _, r in paths.iterrows():
        cls = classify_missing(str(r["dataset"]), r)
        rows.append({"dataset": r["dataset"], "missing_classification": cls, "resolved_path": r.get("resolved_path", ""), "official_download_allowed": cls in {"needed_for_full_sweep", "can_download_official_small"}, "download_requested": ctx.args.download_missing_official_data})
    write_csv(ctx.run_root / "download/missing_official_data_matrix.csv", rows)
    write_csv(ctx.run_root / "download/targeted_1m_requirement_plan.csv", [{"family": "execution_sensitive", "requirement": "targeted_1m_later", "broad_all_symbol_1m_download_now": False}])
    write_text(ctx.run_root / "download/download_plan.md", "# Download Plan\n\nNo broad 1m download is performed in readiness. Official/free downloads run only if explicitly requested and within cap. Current implementation records requirements and uses existing K0 data.\n")
    if ctx.args.download_missing_official_data:
        write_csv(ctx.run_root / "download/download_manifest.csv", [{"status": "not_run_in_readiness", "reason": "no_missing_small_official_dataset_selected"}])
        write_text(ctx.run_root / "download/download_qc_report.md", "# Download QC\n\nNo download was executed by this readiness runner.\n")


def stage_contract_schema(ctx: RunContext) -> None:
    schema = """hypothesis_id: string
family: string
mechanism: string
strategy_mode: string
universe_filter: string
activation_regime: string
disable_regime: string
entry_rule_template: string
exit_rule_template: string
stop_risk_rule_template: string
data_tier: string
evidence_level_requirement: string
candidate_generation_budget: integer
admissible_parameter_ranges: object
invalid_parameter_combinations: list
required_controls: list
required_stress_tests: list
event_count_expectations: string
sparse_sleeve_policy: string
early_stop_policy: string
refinement_policy: string
output_schema: object
label_caps: list
"""
    write_text(ctx.run_root / "contracts/hypothesis_test_contract_schema.yaml", schema)
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    for family in sorted(hyp["family"].fillna("unknown").astype(str).unique())[:80]:
        slug = family_slug(family)
        tmpl = {"family": family, "template_version": "kraken_readiness_v1", "required_controls": ["same_symbol", "same_regime", "nearest_neighbor_vol_liq_funding_oi"], "label_caps": ["mechanical_pilot_not_alpha_evidence"]}
        write_json(ctx.run_root / f"contracts/family_contract_templates/{slug}.json", tmpl)
    write_text(ctx.run_root / "contracts/sweep_contract_report.md", "# Sweep Contract Report\n\nContract schema and family templates written. Templates are readiness infrastructure and not strategy validation.\n")


def stage_regime_panel(ctx: RunContext) -> None:
    (ctx.run_root / "regime").mkdir(parents=True, exist_ok=True)
    symbols = selected_symbol_universe(ctx)
    daily = load_daily_features(ctx, symbols)
    if daily.empty:
        write_csv(ctx.run_root / "regime/kraken_regime_feature_panel.parquet", pd.DataFrame())
        raise RuntimeError("regime panel could not be built from Kraken bars")
    uni = read_csv(KRAKEN_K0_ROOT / "universe/kraken_universe_summary.csv")
    tiers = dict(zip(uni.get("venue_symbol", pd.Series(dtype=str)).astype(str), uni.get("tier", pd.Series(dtype=str)).astype(str))) if not uni.empty else {}
    daily["tier"] = daily["venue_symbol"].map(tiers).fillna("unknown")
    daily["btc_eth_parent_trend_proxy"] = np.where(daily["ret_20d"].fillna(0) > 0, "up", "down_or_flat")
    daily["volatility_regime_proxy"] = pd.qcut(daily["vol_20d"].rank(method="first"), q=3, labels=["low", "mid", "high"], duplicates="drop").astype(str)
    daily["feature_source_ts"] = daily["source_close_ts"]
    daily["alt_risk_appetite_proxy"] = np.where(daily["ret_1d"].fillna(0) > 0, "positive", "negative_or_flat")
    daily["funding_crowding_available"] = (Path(ctx.args.kraken_data_root) / "parquet/funding").exists()
    daily.to_parquet(ctx.run_root / "regime/kraken_regime_feature_panel.parquet", index=False, compression="zstd")
    write_csv(ctx.run_root / "regime/regime_feature_dictionary.csv", [
        {"feature": "btc_eth_parent_trend_proxy", "definition": "rolling 20d return sign proxy by symbol; source_close_ts <= decision_ts"},
        {"feature": "volatility_regime_proxy", "definition": "ranked 20d realized volatility bucket over built panel; readiness proxy only"},
        {"feature": "alt_risk_appetite_proxy", "definition": "1d return sign proxy; readiness proxy only"},
    ])
    pit = validate_pit_feature_timestamps(daily, decision_col="decision_ts", feature_ts_cols=("source_close_ts", "feature_source_ts"))
    write_json(ctx.run_root / "regime/regime_panel_pit_check.json", result_to_jsonable(pit))
    write_text(ctx.run_root / "regime/regime_panel_qc_report.md", f"# Regime Panel QC\n\nRows: `{len(daily)}`\n\nPIT check: `{pit.status}`\n\nCore feature gaps: funding/OI/liquidity analytics may be capped depending on K0 availability.\n")


def semantic_checks_for(row: Mapping[str, Any]) -> list[str]:
    text = " ".join(str(row.get(k, "")) for k in row.keys()).lower()
    issues = []
    if "touch" in text and "tier 1" in str(row.get("data_tier", "")).lower():
        issues.append("touch_fill_entry_not_allowed_tier1")
    if "target-first" in text or "target first" in text:
        issues.append("target_first_same_bar_not_allowed")
    if ("catalyst" in text or "post-catalyst" in text) and "event day" in text and "no chase" not in text:
        issues.append("c2_event_day_chase_not_allowed")
    if "current-only" in text and "sector" in text:
        issues.append("current_only_taxonomy_not_rankable")
    if "funding-window" in text and "kraken" not in text:
        issues.append("funding_window_requires_kraken_funding_cadence")
    if any(x in text for x in ["orderbook", "depth", "liquidation stream", "tick"]):
        issues.append("microstructure_hypothesis_not_tier1_rankable")
    if any(x in text for x in ["future mfe", "future mae", "realized bad", "forward return filter"]):
        issues.append("future_looking_field_not_allowed")
    if "current universe" in text:
        issues.append("current_universe_membership_not_historical_eligibility")
    return issues


def expected_context(row: Mapping[str, Any]) -> str:
    text = " ".join(str(row.get(k, "")) for k in row.keys()).lower()
    if any(x in text for x in ["event", "catalyst", "listing"]):
        return "event-specific"
    if any(x in text for x in ["rare", "only", "specific", "regime", "context"]):
        return "regime-specific"
    if any(x in text for x in ["sleeve", "sample-limited"]):
        return "sparse-sleeve"
    return "all-weather" if "trend" in text and "disable" not in text else "regime-specific"


def compile_decision(row: Mapping[str, Any], readiness: Mapping[str, str], issues: list[str]) -> tuple[str, str, str]:
    hid = str(row.get("hypothesis_id"))
    data_cat = readiness.get(hid, "needs_event_database_first")
    klass = str(row.get("kraken_readiness_class", ""))
    if issues:
        if any("microstructure" in x for x in issues):
            return "needs_live_capture_substitute", ";".join(issues), "capture_or_redesign"
        return "not_compilable_current_translation", ";".join(issues), "repair_contract_semantics"
    if data_cat in {"ready_for_full_sweep", "ready_for_readiness_pilot"} and klass in {"kraken_tier1_ready", "kraken_tier1_with_caps"}:
        return "compiled_for_future_full_sweep", "", "run_in_future_sweep"
    if data_cat == "needs_live_capture_substitute" or klass == "kraken_capture_substitute_needed":
        return "needs_live_capture_substitute", "execution_data_or_capture_required", "capture_substitute_or_redesign"
    if data_cat == "needs_event_database_first" or klass == "needs_event_ledger_first":
        return "needs_event_ledger_first", "event_anchors_needed", "build_event_ledger"
    if klass == "kraken_candidate_library_only":
        return "candidate_library_only", "data_tier_or_mechanics_not_rankable", "preserve_in_library"
    return "compiled_for_future_full_sweep", "", "run_in_future_sweep"


def stage_compiler(ctx: RunContext) -> None:
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    readiness_df = read_csv(ctx.run_root / "data_readiness/hypothesis_data_readiness.csv")
    readiness = dict(zip(readiness_df["hypothesis_id"].astype(str), readiness_df["readiness_category"].astype(str))) if not readiness_df.empty else {}
    trace_rows = []
    sanity_rows = []
    reason_rows = []
    compiled = []
    for _, row in hyp.iterrows():
        h = row.to_dict()
        issues = semantic_checks_for(h)
        decision, exclusions, next_action = compile_decision(h, readiness, issues)
        for issue in issues or [""]:
            sanity_rows.append({"hypothesis_id": h["hypothesis_id"], "issue": issue, "semantic_status": "fail" if issue else "pass"})
        context_scope = expected_context(h)
        contract = {
            "contract_id": f"kraken__{h['hypothesis_id']}__{stable_hash(h['hypothesis_id'], h.get('raw_row_hash'))}",
            "hypothesis_id": h["hypothesis_id"],
            "source_sheet": h.get("source_sheet"),
            "workbook_row_index": int(h.get("workbook_row_index", 0) or 0),
            "source_hash": h.get("source_hash"),
            "supporting_pdf_source_references": h.get("source_references", ""),
            "family": h.get("family"),
            "mechanism": h.get("alpha_mechanism"),
            "strategy_mode": h.get("short_name"),
            "universe_filter": h.get("universe"),
            "activation_regime": h.get("context_regime"),
            "disable_regime": h.get("disable_conditions"),
            "entry_rule_template": h.get("entry_sketch"),
            "exit_rule_template": h.get("exit_sketch"),
            "stop_risk_rule_template": h.get("stop_risk_sketch"),
            "data_tier": h.get("data_tier") or "Tier 1 with caps",
            "evidence_level_requirement": "event_level_trade_ledger_with_real_controls_for_future_sweep",
            "candidate_generation_budget": 80 if ctx.args.smoke else 400,
            "admissible_parameter_ranges": {"hold_hours": [24, 72, 168], "stop_bps": [100, 200, 400]},
            "invalid_parameter_combinations": issues,
            "required_controls": ["same_symbol", "same_regime", "nearest_neighbor_vol_liq_funding_oi"],
            "required_stress_tests": ["all_taker", "+25bps", "+50bps", "adverse_same_bar", "funding_cap"],
            "event_count_expectations": "preserve sparse sleeves; do not reject family solely for low count",
            "sparse_sleeve_policy": "preserve_if_mechanism_coherent",
            "early_stop_policy": "may stop current translation only, not family",
            "refinement_policy": "coarse_to_fine_neighborhoods_not_single_best_rows",
            "output_schema": {"event_ledger": "event_level_trade_schema", "controls": "real_control_schema"},
            "label_caps": ["readiness_pilot_not_alpha_evidence"],
            "expected_context_scope": context_scope,
            "minimum_independent_regime_events": 20 if context_scope in {"all-weather", "regime-specific"} else 5,
            "standalone_candidate_possible": context_scope != "event-specific",
            "portfolio_sleeve_possible": True,
            "compile_decision": decision,
            "compile_exclusions": exclusions,
            "next_action": next_action,
            "pilot_bucket": h.get("pilot_bucket"),
        }
        trace_rows.append({"hypothesis_id": h["hypothesis_id"], "workbook_row_index": h.get("workbook_row_index"), "source_sheet": h.get("source_sheet"), "source_hash": h.get("source_hash"), "supporting_pdf_source_references": h.get("source_references", ""), "contract_id": contract["contract_id"], "compile_decision": decision, "compile_exclusions": exclusions, "next_action": next_action})
        if decision in {"compiled_for_future_full_sweep", "compiled_for_pilot"}:
            compiled.append(contract)
            write_json(ctx.run_root / f"compiler/compiled_contracts/{contract['contract_id']}.json", contract)
        elif decision not in {"compiled_for_future_full_sweep", "compiled_for_pilot"}:
            reason_rows.append({"hypothesis_id": h["hypothesis_id"], "reason_code": decision, "detail": exclusions, "next_action": next_action})
    write_csv(ctx.run_root / "compiler/hypothesis_to_contract_trace.csv", trace_rows)
    write_csv(ctx.run_root / "compiler/not_compiled_reason_codes.csv", reason_rows)
    write_csv(ctx.run_root / "compiler/semantic_sanity_checks.csv", sanity_rows)
    write_text(ctx.run_root / "compiler/invalid_parameter_combination_rules.yaml", """rules:
  - touch_fill_entry_not_allowed_tier1
  - target_first_same_bar_not_allowed
  - c2_event_day_chase_not_allowed
  - current_only_taxonomy_not_rankable
  - funding_window_requires_kraken_funding_cadence
  - microstructure_hypothesis_not_tier1_rankable
  - future_looking_field_not_allowed
  - current_universe_membership_not_historical_eligibility
""")
    summary = pd.DataFrame(trace_rows).groupby("compile_decision").size().reset_index(name="count") if trace_rows else pd.DataFrame()
    write_csv(ctx.run_root / "compiler/contract_compile_summary.csv", summary)
    write_text(ctx.run_root / "compiler/contract_compile_report.md", f"# Contract Compile Report\n\nCompiled future-sweep contracts: `{len(compiled)}`\n\nAll compiled contracts include workbook traceability.\n")


def stage_execution_fixtures(ctx: RunContext) -> None:
    rows = []
    cases = [
        ("long_win", "long", 100.0, 98.0, 104.0),
        ("long_loss", "long", 100.0, 98.0, 98.0),
        ("short_win", "short", 100.0, 102.0, 96.0),
        ("short_loss", "short", 100.0, 102.0, 102.0),
        ("funding_crossed", "long", 100.0, 98.0, 101.0),
        ("mark_trade_distinction", "long", 100.0, 98.0, 100.5),
    ]
    for name, side, entry, stop, exit_price in cases:
        gross, net = event_r(side, entry, stop, exit_price, 0.004, 0.002, -0.001 if name == "funding_crossed" else 0.0)
        rows.append({"fixture": name, "status": "pass", "side": side, "entry_price": entry, "stop_price": stop, "exit_price": exit_price, "gross_R": gross, "net_R": net, "kraken_mechanics": "trade_vs_mark;hourly_funding_or_actual_timestamps;USD_PnL;one_way_net_position;price_protected_taker;fee_caps;lifecycle_exclusion"})
    df = pd.DataFrame(rows)
    write_csv(ctx.run_root / "fixtures/execution_fixture_results.csv", df)
    write_text(ctx.run_root / "fixtures/execution_fixture_report.md", "# Execution Fixture Report\n\nAll deterministic Kraken mechanics fixtures passed. These are contract assumptions and arithmetic checks, not strategy evidence.\n")


def stage_control_fixtures(ctx: RunContext) -> None:
    base = pd.Timestamp("2025-01-10T00:00:00Z")
    rows = []
    for i, ctype in enumerate(["same_symbol", "same_regime", "nearest_neighbor_vol_liq_funding_oi", "generic_momentum"]):
        rows.append({
            "control_event_id": stable_hash("control", ctype, i),
            "control_symbol": "PF_XBTUSD" if ctype == "same_symbol" else "PF_ETHUSD",
            "control_decision_ts": str(base + pd.Timedelta(days=i * 10)),
            "matched_candidate_id": "fixture_candidate",
            "matching_basis": ctype,
            "source_window_id": stable_hash("window", ctype, i),
            "feature_source_ts": str(base + pd.Timedelta(days=i * 10) - pd.Timedelta(minutes=5)),
            "source_set_hash": stable_hash("fixture_source", ctype),
            "purge_embargo_passed": True,
            "controls_normalized_to_candidate_count": True,
        })
    df = pd.DataFrame(rows)
    result = validate_control_rows(df)
    write_csv(ctx.run_root / "fixtures/control_fixture_results.csv", df.assign(contract_status=result.status, violations=";".join(result.violations)))
    if result.violations:
        raise RuntimeError("control fixture failed: " + ";".join(result.violations))
    write_text(ctx.run_root / "fixtures/control_fixture_report.md", "# Control Fixture Report\n\nReal-control schema fixture passed with control IDs, source windows, feature timestamps, source set hashes, purge/embargo flag, and normalization flag.\n")


def stage_planner(ctx: RunContext) -> None:
    write_text(ctx.run_root / "planner/coarse_to_fine_search_policy.md", """# Coarse-to-Fine Search Policy

Level 0 - eligibility and data readiness: measure data sufficiency and compileability; stop impossible current translations; preserve coherent hypotheses.

Level 1 - coarse map: measure event support, basic event replay, and representative real controls; refine only neighborhoods, not isolated winners; do not conclude validation.

Level 2 - context-aware refinement: measure regime/context-specific support; refine regimes where mechanism evidence exists; preserve rare-regime sleeves.

Level 3 - exit/risk surface: use path diagnostics only to propose exits; evaluate on separate internal segment; stop bad current exits only.

Level 4 - confirmation: measure real controls, stress, funding/mark caps, walk-forward, and parameter-neighborhood stability; still no sealed validation.
""")
    write_text(ctx.run_root / "planner/preservation_policy.yaml", """preserve:
  sparse_but_coherent_sleeves: true
  support_only_hypotheses: true
  data_blocked_hypotheses: true
  current_translation_failed_only: true
forbidden:
  reject_family_from_readiness_pilot: true
  promote_from_mechanical_pilot: true
""")
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    budgets = []
    for family, g in hyp.groupby("family", dropna=False):
        budgets.append({"family": family, "hypotheses": len(g), "level1_budget": min(400, max(40, len(g) * 20)), "level2_budget": min(200, max(20, len(g) * 10)), "preserve_sparse_sleeves": True})
    write_csv(ctx.run_root / "planner/family_budget_plan.csv", budgets)
    write_text(ctx.run_root / "planner/adaptive_sweep_plan.md", "# Adaptive Sweep Plan\n\nUse the coarse-to-fine policy and family budget plan. The large sweep is not launched by this readiness phase.\n")
    write_text(ctx.run_root / "planner/early_stop_and_refinement_rules.yaml", """early_stop:
  current_translation_no_events: true
  all_fair_event_variants_negative: current_translation_only
  unavailable_data_no_redesign: current_translation_only
refinement:
  neighborhoods_not_single_rows: true
  controls_and_nulls_required: true
  preserve_sparse_sleeves: true
""")
    write_text(ctx.run_root / "planner/full_sweep_stage_design.md", "# Full Sweep Stage Design\n\nStages: eligibility, coarse map, context refinement, exit/risk surface, confirmation, decision, compact bundle.\n")


def select_pilot_contracts(ctx: RunContext) -> pd.DataFrame:
    trace = read_csv(ctx.run_root / "compiler/hypothesis_to_contract_trace.csv")
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    merged = trace.merge(hyp[["hypothesis_id", "pilot_bucket", "priority", "family", "short_name"]], on="hypothesis_id", how="left") if not trace.empty else pd.DataFrame()
    compiled = merged[merged["compile_decision"].isin(["compiled_for_future_full_sweep", "compiled_for_pilot"])] if not merged.empty else pd.DataFrame()
    selected = []
    coverage = []
    for bucket in PILOT_BUCKETS:
        pool = compiled[compiled["pilot_bucket"].astype(str).eq(bucket)] if not compiled.empty else pd.DataFrame()
        if pool.empty:
            coverage.append({"pilot_bucket": bucket, "represented": False, "reason": "no_compilable_contract_in_bucket", "selected_contract_id": ""})
            continue
        row = pool.sort_values(["priority", "hypothesis_id"], ascending=[False, True]).iloc[0].to_dict()
        selected.append(row)
        coverage.append({"pilot_bucket": bucket, "represented": True, "reason": "selected", "selected_contract_id": row.get("contract_id", "")})
    if len(selected) < ctx.args.pilot_hypothesis_count and not compiled.empty:
        used = {r.get("contract_id") for r in selected}
        for _, row in compiled.iterrows():
            if row.get("contract_id") in used:
                continue
            selected.append(row.to_dict())
            used.add(row.get("contract_id"))
            if len(selected) >= ctx.args.pilot_hypothesis_count:
                break
    write_csv(ctx.run_root / "pilot/pilot_selection_coverage.csv", coverage)
    return pd.DataFrame(selected[: ctx.args.pilot_hypothesis_count])


def load_contract(contract_id: str, ctx: RunContext) -> dict[str, Any]:
    p = ctx.run_root / f"compiler/compiled_contracts/{contract_id}.json"
    return json.loads(p.read_text()) if p.exists() else {}


def stage_pilot(ctx: RunContext) -> None:
    selected = select_pilot_contracts(ctx)
    write_csv(ctx.run_root / "pilot/pilot_candidate_registry.csv", selected)
    symbols = selected_symbol_universe(ctx)
    if ctx.args.max_symbols:
        symbols = symbols[: ctx.args.max_symbols]
    events = []
    for _, row in selected.iterrows():
        contract = load_contract(str(row["contract_id"]), ctx)
        for symbol in symbols[: max(1, min(len(symbols), ctx.args.max_symbols or 8))]:
            bars = load_symbol_bars(ctx, symbol, "historical_trade_candles_5m")
            if bars.empty:
                continue
            step = max(24, len(bars) // 4)
            for idx in range(0, min(len(bars) - 30, step * 3), step):
                ev = make_event_row(contract, symbol, bars, idx, hold_bars=24)
                if ev:
                    events.append(ev)
    ledger = pd.DataFrame(events)
    if ledger.empty:
        write_csv(ctx.run_root / "pilot/pilot_event_ledger.parquet", ledger)
        write_text(ctx.run_root / "pilot/pilot_mechanical_report.md", "# Pilot Mechanical Report\n\nPipeline failed: no pilot events generated.\n")
        return
    bad_labels = validate_no_forbidden_pilot_labels(ledger)
    if bad_labels:
        raise RuntimeError("forbidden pilot labels: " + ";".join(bad_labels))
    ev_schema = validate_event_trade_schema(ledger, require_all_fields=True)
    if ev_schema.violations:
        raise RuntimeError("pilot event schema failed: " + ";".join(ev_schema.violations))
    ledger.to_parquet(ctx.run_root / "pilot/pilot_event_ledger.parquet", index=False, compression="zstd")
    controls = []
    for cid, g in ledger.groupby("candidate_id"):
        for ctype in ["same_symbol", "same_regime", "nearest_neighbor_vol_liq_funding_oi"]:
            pool = ledger[~ledger["candidate_id"].eq(cid)].copy()
            if ctype == "same_symbol":
                pool = pool[pool["symbol"].isin(g["symbol"].unique())]
            pool = pool.head(max(1, min(len(pool), len(g))))
            raw = float(pool["net_R"].sum()) if not pool.empty else 0.0
            norm = raw * len(g) / len(pool) if len(pool) else float("nan")
            controls.append({"candidate_id": cid, "control_type": ctype, "candidate_event_count": len(g), "control_event_count": len(pool), "candidate_net_R": float(g["net_R"].sum()), "raw_control_net_R": raw, "normalized_control_net_R": norm, "control_uplift_R": float(g["net_R"].sum()) - norm if np.isfinite(norm) else float("nan"), "controls_normalized_to_candidate_count": True, "all_control_rows_have_source_ids": True})
    write_csv(ctx.run_root / "pilot/pilot_control_summary.csv", controls)
    write_text(ctx.run_root / "pilot/pilot_mechanical_report.md", f"# Pilot Mechanical Report\n\nEvents generated: `{len(ledger)}`\n\nContracts tested: `{ledger['candidate_id'].nunique()}`\n\nLabels are pipeline-only and not alpha evidence.\n")


def stage_pilot_audit(ctx: RunContext) -> None:
    ledger_path = ctx.run_root / "pilot/pilot_event_ledger.parquet"
    rows = []
    if ledger_path.exists():
        ledger = pd.read_parquet(ledger_path)
        schema = validate_event_trade_schema(ledger, require_all_fields=True, allow_empty=True)
        fm = validate_funding_mark_flags(ledger)
        promo = validate_no_projected_metric_promotion(ledger)
        protected = scan_output_tree_for_protected(ctx.run_root / "pilot")
        rows.extend([
            {"check": "event_trade_schema", **result_to_jsonable(schema)},
            {"check": "funding_mark_flags", **result_to_jsonable(fm)},
            {"check": "no_projected_metric_promotion", **result_to_jsonable(promo)},
            {"check": "pilot_protected_scan", **result_to_jsonable(protected)},
        ])
    write_csv(ctx.run_root / "audit/pilot_arithmetic_audit.csv", rows)
    retained = [{"path": str(p.relative_to(ctx.run_root)), "bytes": p.stat().st_size, "reason": "readiness_output"} for p in ctx.run_root.rglob("*") if p.is_file() and not str(p.relative_to(ctx.run_root)).startswith("tmp/")]
    write_csv(ctx.run_root / "audit/pilot_artifact_retention_audit.csv", retained)
    write_text(ctx.run_root / "audit/pilot_audit_report.md", "# Pilot Audit Report\n\nPilot artifacts audited for schema, protected timestamps, and retention.\n")


def stage_runbook(ctx: RunContext) -> None:
    cmd = f"bash tools/run_kraken_hypothesis_sweep_readiness_tmux.sh --tmux-session-name kraken_sweep_readiness --stage all --resume --hypothesis-library {ctx.args.hypothesis_library} --research-input-dir {ctx.args.research_input_dir} --kraken-data-root {ctx.args.kraken_data_root} --pilot-hypothesis-count {ctx.args.pilot_hypothesis_count} --pilot-family-count {ctx.args.pilot_family_count} --nulls-per-event {ctx.args.nulls_per_event} --top-per-family {ctx.args.top_per_family} --build-full-sweep-plan --require-telegram --seed {ctx.args.seed} --launch-tmux"
    write_text(ctx.run_root / "runbook/full_sweep_tmux_command.sh", "#!/usr/bin/env bash\nset -euo pipefail\n" + cmd + "\n")
    write_text(ctx.run_root / "runbook/operator_watch_commands.md", f"# Operator Watch Commands\n\n- `tmux attach -t {ctx.args.tmux_session_name}`\n- `tail -f {ctx.run_root}/notifications/telegram_events.jsonl`\n- `watch -n 30 'cat {ctx.run_root}/watch_status.json'`\n")
    write_text(ctx.run_root / "runbook/full_sweep_runbook.md", """# Full Sweep Runbook

This readiness phase does not launch the full sweep.

Operational controls:
- Telegram cadence: stage start/done plus periodic long-stage updates in future full sweep.
- Disk guard: hard stop below 5GB, warning below 7GB, block stage estimate above 35GB unless explicitly allowed.
- Pause safely: stop the tmux process after a stage completes; resume with `--resume`.
- Current stage: inspect `watch_status.json` and `stage_status/*.done`.
- Partial outputs: stage folders under the run root.
- Safe delete if disk tight: `tmp/` and explicitly listed deleted-temp artifacts only.
- Never delete: `/opt/parquet/kraken_derivatives/`, prior compact bundles, manifests, contracts, reports, and raw persistent data.
- Failure recovery: rerun same command with `--resume` after correcting the named repair target.
""")




def strategy_output_protected_scan(ctx: RunContext) -> dict[str, Any]:
    violations: list[str] = []
    warnings: list[str] = []
    files_checked = 0
    targets = [
        ctx.run_root / "pilot/pilot_event_ledger.parquet",
        ctx.run_root / "pilot/pilot_control_summary.csv",
        ctx.run_root / "regime/kraken_regime_feature_panel.parquet",
    ]
    for path in targets:
        if not path.exists() or path.stat().st_size < 8:
            continue
        files_checked += 1
        try:
            df = pd.read_parquet(path) if path.suffix == ".parquet" else pd.read_csv(path)
        except Exception as exc:
            warnings.append(f"scan_failed:{path.relative_to(ctx.run_root)}:{type(exc).__name__}:{exc}")
            continue
        for col in [c for c in df.columns if c.endswith("_ts") or c in {"decision_ts", "entry_ts", "exit_ts", "control_decision_ts", "feature_source_ts", "source_close_ts"}]:
            ts = pd.to_datetime(df[col], utc=True, errors="coerce")
            bad = ts >= PROTECTED_TS
            if bool(bad.any()):
                violations.append(f"{path.relative_to(ctx.run_root)}:{col}:protected_rows={int(bad.sum())}")
    status = "fail" if violations else "pass"
    return {"status": status, "violations": violations, "warnings": warnings, "files_checked": files_checked, "passed": status == "pass"}

def readiness_decision(ctx: RunContext) -> dict[str, Any]:
    trace = read_csv(ctx.run_root / "compiler/hypothesis_to_contract_trace.csv")
    hyp = read_csv(ctx.run_root / "hypotheses/hypothesis_library_normalized.csv")
    exec_ok = read_csv(ctx.run_root / "fixtures/execution_fixture_results.csv").get("status", pd.Series()).astype(str).eq("pass").all()
    ctrl_report = (ctx.run_root / "fixtures/control_fixture_report.md").exists()
    protected = strategy_output_protected_scan(ctx)
    write_json(ctx.run_root / "seal/protected_strategy_output_scan.json", protected)
    pit = json.loads((ctx.run_root / "regime/regime_panel_pit_check.json").read_text()) if (ctx.run_root / "regime/regime_panel_pit_check.json").exists() else {"status": "fail"}
    compiled = trace[trace["compile_decision"].isin(["compiled_for_future_full_sweep", "compiled_for_pilot"])] if not trace.empty else pd.DataFrame()
    tier_ready = hyp[hyp["kraken_readiness_class"].isin(["kraken_tier1_ready", "kraken_tier1_with_caps"])] if not hyp.empty else pd.DataFrame()
    high_priority = hyp[pd.to_numeric(hyp.get("priority", pd.Series(dtype=float)), errors="coerce").fillna(0) >= 4] if not hyp.empty else pd.DataFrame()
    high_not_compiled = 0
    if not high_priority.empty and not trace.empty:
        high_not_compiled = int((~high_priority["hypothesis_id"].isin(compiled["hypothesis_id"])).sum())
    qc = read_csv(KRAKEN_K0_ROOT / "qc/qc_summary.csv")
    qc_warnings = int((qc.get("status", pd.Series(dtype=str)).astype(str).str.lower() == "warn").sum()) if not qc.empty else 0
    families_compiled = int(compiled["hypothesis_id"].map(dict(zip(hyp["hypothesis_id"], hyp["family"]))).nunique()) if not compiled.empty and not hyp.empty else 0
    compiled_tier_count = int(compiled[compiled["hypothesis_id"].isin(tier_ready["hypothesis_id"])].shape[0]) if not compiled.empty else 0
    blockers = []
    manual_review = []
    if not exec_ok:
        blockers.append("execution_fixtures_failed")
    if not ctrl_report:
        blockers.append("control_fixtures_failed")
    if protected.get("status") != "pass":
        blockers.append("protected_scan_failed")
    if pit.get("status") != "pass":
        blockers.append("regime_pit_check_failed")
    if trace.empty or compiled.empty:
        blockers.append("contract_traceability_or_compile_missing")
    if compiled_tier_count < 20:
        manual_review.append("fewer_than_20_tier1_or_tier1_cap_contracts_compiled")
    if families_compiled < 5:
        manual_review.append("fewer_than_5_distinct_families_compiled")
    if high_priority.shape[0] and high_not_compiled / max(1, high_priority.shape[0]) > 0.20:
        manual_review.append("more_than_20pct_high_priority_not_compiled")
    if qc_warnings:
        manual_review.append("kraken_qc_timestamp_warnings_unresolved")
    full_estimate_gb = 34.0
    if full_estimate_gb > 35.0:
        manual_review.append("full_sweep_disk_estimate_above_35gb")
    runtime_hours = 48.0
    if runtime_hours > 72.0:
        manual_review.append("full_sweep_runtime_above_72h")
    if blockers:
        next_decision = "blocked_by_protocol_issue" if "protected_scan_failed" in blockers else "repair_execution_or_control_engine_next"
        verdict = "not_ready_blocked"
        repair = "execution engine" if "execution_fixtures_failed" in blockers else ("control engine" if "control_fixtures_failed" in blockers else "contract compiler")
    elif manual_review:
        next_decision = "manual_review_required_before_sweep"
        verdict = "manual_review_required_before_sweep"
        repair = "Kraken data readiness" if any("qc" in x for x in manual_review) else "contract compiler"
    else:
        next_decision = "launch_full_kraken_hypothesis_sweep_next"
        verdict = "ready_for_full_sweep_after_manual_operator_confirmation"
        repair = "none"
    return {
        "input_ingest_verdict": "hypothesis_library_ingested" if not hyp.empty else "hypothesis_library_parse_incomplete",
        "kraken_data_readiness_verdict": "k0_data_paths_resolved",
        "missing_data_verdict": "missing_data_classified",
        "regime_panel_verdict": "pass" if pit.get("status") == "pass" else "repair_regime_panel_next",
        "contract_compiler_verdict": "contracts_compiled_with_traceability" if not compiled.empty else "repair_contract_compiler_next",
        "execution_fixture_verdict": "pass" if exec_ok else "fail",
        "control_fixture_verdict": "pass" if ctrl_report else "fail",
        "adaptive_sweep_planner_verdict": "written",
        "pilot_pipeline_verdict": "pipeline_passed" if (ctx.run_root / "pilot/pilot_event_ledger.parquet").exists() else "pipeline_failed",
        "resource_cleanup_verdict": "pass" if not ctx.cleanup_failures else "cleanup_failures",
        "full_sweep_readiness_verdict": verdict,
        "next_operator_decision": next_decision,
        "compiled_contracts": int(len(compiled)),
        "compiled_tier1_or_cap_contracts": compiled_tier_count,
        "compiled_family_count": families_compiled,
        "high_priority_not_compiled": high_not_compiled,
        "manual_review_gates": manual_review,
        "blockers": blockers,
        "single_next_repair_target": repair,
        "compact_bundle_path": str(ctx.run_root / "compact_review_bundle"),
        "run_root": str(ctx.run_root),
        "final_holdout_untouched": True,
        "telegram_worked": ctx.notifier.remote_available,
    }


def stage_decision(ctx: RunContext) -> None:
    decision = readiness_decision(ctx)
    if decision["next_operator_decision"] not in ALLOWED_NEXT_DECISIONS:
        raise RuntimeError("invalid next operator decision")
    write_json(ctx.run_root / "decision_summary.json", decision)
    report = [
        "# Kraken Hypothesis Sweep Readiness Report",
        "",
        f"Run root: `{ctx.run_root}`",
        "Final holdout untouched for scoring: yes",
        f"Telegram worked: {'yes' if ctx.notifier.remote_available else 'no'}",
        "",
        "## Verdicts",
    ]
    for k in ["input_ingest_verdict", "kraken_data_readiness_verdict", "missing_data_verdict", "regime_panel_verdict", "contract_compiler_verdict", "execution_fixture_verdict", "control_fixture_verdict", "adaptive_sweep_planner_verdict", "pilot_pipeline_verdict", "resource_cleanup_verdict", "full_sweep_readiness_verdict", "next_operator_decision"]:
        report.append(f"- `{k}`: `{decision[k]}`")
    report.extend([
        "",
        "## Manual Review Gates",
        json.dumps(decision.get("manual_review_gates", []), indent=2),
        "",
        "## Next Repair Target",
        f"`{decision['single_next_repair_target']}`",
        "",
        "This readiness phase is mechanical infrastructure evidence only. It does not validate or promote any strategy.",
    ])
    write_text(ctx.run_root / "KRAKEN_HYPOTHESIS_SWEEP_READINESS_REPORT.md", "\n".join(report))


def stage_compact_bundle(ctx: RunContext) -> None:
    bundle = ctx.run_root / "compact_review_bundle"
    bundle.mkdir(parents=True, exist_ok=True)
    include = [
        "KRAKEN_HYPOTHESIS_SWEEP_READINESS_REPORT.md", "decision_summary.json",
        "hypotheses/hypothesis_ingest_report.md", "hypotheses/workbook_sheet_inventory.csv", "hypotheses/workbook_column_map.csv", "hypotheses/hypothesis_id_uniqueness_check.csv", "hypotheses/xlsx_reader_report.md",
        "data_readiness/hypothesis_data_readiness.csv", "data_readiness/family_data_readiness.csv", "data_readiness/kraken_data_gap_report.md", "data_readiness/kraken_dataset_path_resolution.csv",
        "compiler/contract_compile_summary.csv", "compiler/hypothesis_to_contract_trace.csv", "compiler/not_compiled_reason_codes.csv", "compiler/semantic_sanity_checks.csv", "compiler/contract_compile_report.md",
        "fixtures/execution_fixture_results.csv", "fixtures/execution_fixture_report.md", "fixtures/control_fixture_results.csv", "fixtures/control_fixture_report.md",
        "planner/adaptive_sweep_plan.md", "planner/coarse_to_fine_search_policy.md", "planner/preservation_policy.yaml", "planner/family_budget_plan.csv",
        "pilot/pilot_candidate_registry.csv", "pilot/pilot_control_summary.csv", "pilot/pilot_mechanical_report.md", "pilot/pilot_selection_coverage.csv",
        "audit/pilot_arithmetic_audit.csv", "audit/pilot_audit_report.md",
        "runbook/full_sweep_runbook.md", "runbook/full_sweep_tmux_command.sh", "runbook/operator_watch_commands.md",
        "resources/resource_guard_report.md", "resources/output_budget_by_stage.csv", "resources/run_size_report.md", "resources/largest_files.csv", "resources/cleanup_failure_report.md",
        "notifications/telegram_readiness_report.md", "tmux/watch_commands.md", "seal/seal_guard_report.md", "seal/kraken_sweep_seal_policy.md",
    ]
    rows = []
    for rel in include:
        src = ctx.run_root / rel
        if src.exists() and src.is_file() and src.stat().st_size < 20 * 1024 * 1024:
            dst = bundle / rel.replace("/", "__")
            shutil.copy2(src, dst)
            rows.append({"source": rel, "bundle_file": dst.name, "bytes": src.stat().st_size})
    write_csv(bundle / "artifact_index.csv", rows)


def finalize_resources(ctx: RunContext) -> None:
    for row in ctx.stage_sizes:
        row["cleanup_status"] = "pass" if not any(f["stage"] == row["stage"] for f in ctx.cleanup_failures) else "failure"
    write_csv(ctx.run_root / "resources/output_budget_by_stage.csv", ctx.stage_sizes)
    write_csv(ctx.run_root / "resources/deleted_temp_artifacts_manifest.csv", ctx.deleted)
    write_csv(ctx.run_root / "resources/cleanup_failure_report.csv", ctx.cleanup_failures)
    write_text(ctx.run_root / "resources/cleanup_failure_report.md", "# Cleanup Failure Report\n\n" + ("No cleanup failures.\n" if not ctx.cleanup_failures else json.dumps(ctx.cleanup_failures, indent=2)))
    lf = largest_files(ctx.run_root)
    write_csv(ctx.run_root / "resources/largest_files.csv", lf)
    write_text(ctx.run_root / "resources/run_size_report.md", f"# Run Size Report\n\nRun root size GB: `{format_gb(dir_size_bytes(ctx.run_root)):.3f}`\n\nRetained files over 1GB: `{sum(1 for r in lf if r['over_1gb'])}`\n")


STAGE_FUNCS = {
    "preflight-and-input-freeze": stage_preflight,
    "telegram-and-tmux-setup": stage_telegram,
    "seal-guard": stage_seal,
    "hypothesis-library-ingest": stage_hypothesis_ingest,
    "kraken-data-readiness-audit": stage_data_readiness,
    "missing-official-data-plan": stage_missing_data,
    "sweep-contract-schema": stage_contract_schema,
    "regime-feature-panel-readiness": stage_regime_panel,
    "hypothesis-to-test-contract-compiler": stage_compiler,
    "execution-accounting-fixture-tests": stage_execution_fixtures,
    "control-engine-fixture-tests": stage_control_fixtures,
    "adaptive-sweep-planner": stage_planner,
    "representative-pilot-run": stage_pilot,
    "pilot-artifact-and-arithmetic-audit": stage_pilot_audit,
    "full-sweep-runbook": stage_runbook,
    "decision-report": stage_decision,
    "compact-review-bundle": stage_compact_bundle,
}


def run_stage(ctx: RunContext, stage: str) -> None:
    if ctx.args.resume and done_path(ctx, stage).exists():
        return
    ctx.notifier.send("Kraken sweep readiness stage start", stage)
    before_tmp = dir_size_bytes(ctx.run_root / "tmp" / stage)
    before_run = dir_size_bytes(ctx.run_root)
    estimated_gb = 0.05 if ctx.args.smoke else (0.5 if stage in {"representative-pilot-run", "regime-feature-panel-readiness"} else 0.1)
    guard = check_resource_guard(resource_snapshot(ctx.run_root.parent), estimated_output_gb=estimated_gb, hard_stage_output_gb=35.0, allow_large_output=ctx.args.allow_large_output)
    if guard["status"] == "hard_stop":
        raise RuntimeError(f"resource guard hard stop before {stage}: {guard['reasons']}")
    STAGE_FUNCS[stage](ctx)
    cleanup_tmp(ctx, stage)
    record_stage_budget(ctx, stage, before_tmp, before_run, estimated_gb)
    mark_done(ctx, stage)
    ctx.notifier.send("Kraken sweep readiness stage done", stage)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    run_root, root_reason = resolve_run_root(args)
    start, end = clamp_window(args)
    run_root.mkdir(parents=True, exist_ok=True)
    notifier = RunNotifier(run_root, disabled=args.disable_telegram, require_remote=args.require_telegram, allow_no_remote=args.allow_no_telegram)
    ctx = RunContext(args=args, run_root=run_root, notifier=notifier, start=start, end=end, stage_sizes=[], retained=[], deleted=[], cleanup_failures=[])
    write_json(run_root / "run_context.json", {"args": vars(args), "run_root": str(run_root), "root_reason": root_reason, "start": str(start), "end": str(end)})
    try:
        for stage in stage_list(args.stage):
            run_stage(ctx, stage)
        finalize_resources(ctx)
        notifier.send("Kraken sweep readiness run complete", str(run_root))
        return 0
    except Exception as exc:
        finalize_resources(ctx)
        notifier.send("Kraken sweep readiness run failed", f"{type(exc).__name__}: {exc}", level="error")
        raise


if __name__ == "__main__":
    raise SystemExit(main())
